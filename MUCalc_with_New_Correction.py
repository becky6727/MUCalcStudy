import os, sys, time
import datetime
import numpy
import argparse
from openpyxl import load_workbook #read excel files
import src.FSF as FSFactor

#options
Parser = argparse.ArgumentParser(description = 'options for analysis')

Parser.add_argument('-E',
                    type = str,
                    dest = 'Energy',
                    default = 'S200',
                    help = 'Energy for analysis')

Parser.add_argument('-l',
                    type = str,
                    default = None,
                    dest = 'List',
                    help = 'List File Name')

Parser.add_argument('-spl',
                    action = 'store_true',
                    dest = 'isSPL',
                    help = 'option flag for spline interpolation')

args = Parser.parse_args()

Energy = args.Energy
List = args.List
isSPL = args.isSPL

#errors
if(List == None):
    print 'please type the list file name with option -l'
    sys.exit()
    pass

f = open(List)
FileExls = f.readlines()
f.close()

#data output
DArray = []

#object for calculating field size factor with field size and nozzle position
ObjFSF = FSFactor.FSF(Energy)

#option for FSF calculation
if(isSPL):
    print 'use spline option for calculating FSF'
    OptFSF = 'spl'
else:
    print 'not use spline option for calculating FSF'
    OptFSF = 'fit'
    pass

for i in range(len(FileExls)):
    
    File = FileExls[i].rstrip()
    
    wb = load_workbook(File, data_only = True)

    for j in range(12):
        
        #initialize
        tmpArray = [0 for i in range(8)]
        QA_ver = -1
        isPDD_Broken = False
        isPar_Broken = False
        
        ws = wb.worksheets[j]

        #is this QA sheet?
        if(str(ws.cell('L2').value) == 'None'):
            continue
        
        #has no contents
        if(ws.cell('B10').value == '#N/A'):
            continue
        
        #has no measured MU
        if(ws.cell('D49').value == '#DIV/0!'):
            continue

        #avoid example sheet
        if(str(ws.cell('B2').value) == 'None'):
            continue

        #has no date information
        #if(str(ws.cell('E2').value) == 'None'):
        #    continue
        
        #QA sheet version check
        if(str(ws.cell('G48').value) == 'Gy'):
            QA_ver = 2
        elif(str(ws.cell('G48').value) == 'None'):
            QA_ver = 3
        else:
            QA_ver = -1
            pass
        
        if(QA_ver < 0):
            continue
        
        if((QA_ver == 2) and (ws.cell('D41').value == '#DIV/0!')):
            continue
        
        tmpDate = (ws.cell('E2').internal_value) #date
                
        #error
        if(str(tmpDate) == '2013/10/'):
            continue
        
        if(str(tmpDate) == 'None'):
            Date = datetime.date(1900, 1, 1)
        elif(isinstance(tmpDate, unicode)):
            Date = datetime.datetime.strptime(str(tmpDate), '%Y.%m.%d')
            Date = Date.date()
        elif(int(tmpDate) > 2958465):
            Date = datetime.datetime.strptime(str(tmpDate), '%Y%m%d')
            Date = Date.date()
        else:
            Date = (datetime.date(1900, 1, 1)) + datetime.timedelta(int(tmpDate) - 2) #serial -> date
            pass
        
        EnergyID = ws.cell('B7').value #energy
        FS = ws.cell('F56').value #mm
        Nozzle = ws.cell('B11').value #mm
        SOBP = ws.cell('B8').value #mm
        RS = ws.cell('B9').value #mm
        
        if(not(str(EnergyID) == Energy)):
            continue
        
        if(QA_ver != 3):
            
            dMU_Meas = (100.0* (ws.cell('D44').value))/(ws.cell('C46').value) 
            #measured dose ---> dose/MU
        
            #calc MU
            ROF = ws.cell('B29').value
            SOBPF = ws.cell('B30').value
            RSF = ws.cell('B31').value
            FSF = ws.cell('B32').value
            CSF = ws.cell('D41').value
            #PlanGy = ws.cell('D36').value
            PlanGy = ws.cell('C46').value
            
            if(str(PlanGy) == 'None'):
                isPar_Broken = True
                break
            
            #dMU_VQA = ROF* SOBPF* RSF* FSF* CSF #calc dose/MU
            dMU_VQA = ROF* SOBPF* RSF* FSF #calc dose/MU
        
        else:
                        
            dMU_Meas = (ws.cell('D47').value) #measured Gy --> dose/MU
            
            if(isinstance(dMU_Meas, unicode)):
                isPar_Broken = True
                break
            
            dMU_Meas = (100.0* dMU_Meas)/50.0
            
            #calc MU
            ROF = ws.cell('B29').value
            SOBPF = ws.cell('B30').value
            RSF = ws.cell('B31').value
            FSF = ws.cell('B32').value
            CSF = ws.cell('B33').value
            PlanGy = ws.cell('D36').value
            
            if(str(PlanGy) == 'None'):
                isPar_Broken = True
                break
            
            #dMU_VQA = ROF* SOBPF* RSF* FSF* CSF #calc dose/MU
            dMU_VQA = ROF* SOBPF* RSF* FSF #calc dose/MU

            pass
        
        if(isPar_Broken):
            continue
        
        #calc modified d/MU with corrected FSF
        FS = FS/10.0 #mm -> cm
        FS_Area = FS**2 #cm -> cm^{2}

        Nozzle = Nozzle/10.0 #mm -> cm
        Nozzle = Nozzle + 26.9 #distance from snout -> MLC-edge

        SOBP = SOBP/10.0 #mm -> cm
        
        FSF_mod = ObjFSF.GetValue(FS_Area, Nozzle, OptFSF)
        
        #dMU_mod = ROF* SOBPF* RSF* CSF* FSF_mod #modified calc dose/MU
        dMU_mod = ROF* SOBPF* RSF* FSF_mod #modified calc dose/MU w/o CSF
        
        tmpArray[0] = int(Date.toordinal()) #date(string) -> date(serial)
        tmpArray[1] = FS #cm
        tmpArray[2] = Nozzle #cm, distance from iso-center to MLC-edge
        tmpArray[3] = SOBP #cm
        tmpArray[4] = RS #range shifter
        tmpArray[5] = dMU_VQA
        tmpArray[6] = dMU_Meas
        tmpArray[7] = dMU_mod

        DArray.append(tmpArray)
        
        pass
    
    pass

#convert data type
DArray = numpy.array(DArray, dtype = float)

#output
if(isSPL):
    Output = './Data/CorrectedMU_wo_CSF_%s_SPL.dat' %(Energy)
else:
    Output = './Data/CorrectedMU_wo_CSF_%s.dat' %(Energy)
    pass

numpy.savetxt(Output, DArray, fmt = '%d %.4f %.2f %.1f %.1f %.4f %.4f %.4f', delimiter = ' ')
