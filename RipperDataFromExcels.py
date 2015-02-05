#coding: utf-8

import os, sys, time, codecs
import datetime
import math
import numpy
import argparse
from openpyxl import load_workbook #read excel files

#options
Parser = argparse.ArgumentParser(description = 'this program works in order to rip the data from Patient QA sheet excels')

Parser.add_argument('-debug',
                    action = 'store_true',
                    dest = 'isDebug',
                    help = 'Flag for debug mode')

Parser.add_argument('-o',
                    type = str,
                    default = None,
                    dest = 'Output',
                    help = 'Output File Name')

Parser.add_argument('-l',
                    type = str,
                    default = None,
                    dest = 'List',
                    help = 'List File Name')

args = Parser.parse_args()

isDebug = args.isDebug
Output = args.Output
List = args.List

#errors
if(Output == None):
    print 'please type the output file name with option -o'
    sys.exit()
    pass

if(List == None):
    print 'please type the list file name with option -l'
    sys.exit()
    pass

#List = 'SampleExcels.list'

f = open(List)
FileExls = f.readlines()
f.close()

#data array
Array = []

#if debug mode ON...
if(isDebug):
    d = datetime.datetime.today()
    DebugLog = './log/DebugLog_%s%s%s_%s_%s_%s.log' %(d.year, d.month, d.day, d.hour, d.minute, d.second)
    print 'Debug Mode is ON ---> log data: %s' %(DebugLog)
    pass

for i in range(len(FileExls)):
#for i in range(2):
    
    File = FileExls[i].rstrip()
    #print File
    
    wb = load_workbook(File, data_only = True)

    for j in range(12):
        
        #initialize
        DArray = [0 for i in range(23)]
        QA_ver = -1
        isPDD_Broken = False
        isPar_Broken = False
        
        ws = wb.worksheets[j]

        #is this QA sheet?
        if(str(ws.cell('L2').value) == 'None'):
            continue
        
        #has no contents
        if(ws.cell('B10').value == '#N/A'):
            continue
        
        #has no measured MU
        if(ws.cell('D49').value == '#DIV/0!'):
            continue

        #avoid example sheet
        if(str(ws.cell('B2').value) == 'None'):
            continue
        
        #QA sheet version check
        if(str(ws.cell('G48').value) == 'Gy'):
            QA_ver = 2
        elif(str(ws.cell('G48').value) == 'None'):
            QA_ver = 3
        else:
            QA_ver = -1
            pass
        
        if(QA_ver < 0):
            continue
        
        if((QA_ver == 2) and (ws.cell('D41').value == '#DIV/0!')):
            continue
        
        if(isDebug):
            LogIN = open(DebugLog, 'a')
            
            ID = str(ws.cell('B2').value)
            SheetName = ws.title
            SheetName = SheetName.encode('utf-8')
            SheetName = str(SheetName)
            
            #File name
            LogIN.write('File Name: ')
            LogIN.write(File)
            LogIN.write(', sheet name: ')
            
            #sheet name
            LogIN.write(SheetName)
            LogIN.write(', QA version: ')
            
            #QA sheet version
            LogIN.write(str(QA_ver))
            LogIN.write('\n')
                        
            LogIN.close()
            pass
        
        DArray[0] = ws.cell('B2').value #id
        DArray[1] = ws.cell('B7').value #energy
        DArray[2] = ws.cell('B8').value #SOBP
        DArray[3] = ws.cell('B9').value #RS
        DArray[4] = ws.cell('B10').value #Range
        DArray[5] = ws.cell('B11').value #Snout position
        DArray[6] = ws.cell('B12').value #Gantry angle
        DArray[7] = ws.cell('B13').value #Normalized depth
        DArray[8] = ws.cell('B16').value #iso position
        DArray[9] = ws.cell('B17').value #distal position
        DArray[10] = ws.cell('B18').value #proxup position
        DArray[11] = ws.cell('B20').value #SOBP center position
        DArray[12] = ws.cell('F56').value #Field Size
        DArray[13] = ws.cell('J16').value #dose at iso
        DArray[14] = ws.cell('J17').value #dose at distal
        DArray[15] = ws.cell('J18').value #dose at proxup
        DArray[19] = ws.cell('H18').value #gamma-index analysis at proxup       
        DArray[22] = ws.cell('D125').value #SAD
        
        if(QA_ver != 3):
            
            #DArray[20] = ws.cell('D46').value #measured MU
            DArray[20] = (100.0* (ws.cell('D44').value))/(ws.cell('C46').value) #measured dose ---> dose/MU
        
            #calc MU
            ROF = ws.cell('B29').value
            SOBPF = ws.cell('B30').value
            RSF = ws.cell('B31').value
            FSF = ws.cell('B32').value
            CSF = ws.cell('D41').value
            #PlanGy = ws.cell('D36').value
            PlanGy = ws.cell('C46').value
            
            if(isDebug):
                LogIN = open(DebugLog, 'a')
                
                #d/MU parameters
                LogIN.write('Parameters: ')
                LogIN.write(str(ROF))
                LogIN.write(' ')
                LogIN.write(str(SOBPF))
                LogIN.write(' ')
                LogIN.write(str(RSF))
                LogIN.write(' ')
                LogIN.write(str(FSF))
                LogIN.write(' ')
                LogIN.write(str(CSF))
                LogIN.write(' ')
                LogIN.write(str(PlanGy))
                LogIN.write('\n')
                
                LogIN.close()
                pass
            
            if(str(PlanGy) == 'None'):
                isPar_Broken = True
                break
            
            #DArray[21] = (100.0* PlanGy)/(ROF* SOBPF* RSF* FSF* CSF) #calc MU
            DArray[21] = ROF* SOBPF* RSF* FSF* CSF #calc dose/MU
        
        else:
                        
            #DArray[20] = ws.cell('D49').value #measured MU
            DArray[20] = (ws.cell('D47').value) #measured Gy --> dose/MU
            
            if(isinstance(DArray[20], unicode)):
                isPar_Broken = True
                break
            
            DArray[20] = (100.0* DArray[20])/50.0
            
            #calc MU
            ROF = ws.cell('B29').value
            SOBPF = ws.cell('B30').value
            RSF = ws.cell('B31').value
            FSF = ws.cell('B32').value
            CSF = ws.cell('B33').value
            PlanGy = ws.cell('D36').value
            
            if(isDebug):
                LogIN = open(DebugLog, 'a')
                
                #d/MU parameters
                LogIN.write('Parameters: ')
                LogIN.write(str(ROF))
                LogIN.write(' ')
                LogIN.write(str(SOBPF))
                LogIN.write(' ')
                LogIN.write(str(RSF))
                LogIN.write(' ')
                LogIN.write(str(FSF))
                LogIN.write(' ')
                LogIN.write(str(CSF))
                LogIN.write(' ')
                LogIN.write(str(PlanGy))
                LogIN.write('\n')
                
                LogIN.close()
                pass
            
            if(str(PlanGy) == 'None'):
                isPar_Broken = True
                break
            
            #DArray[21] = (100.0* PlanGy)/(ROF* SOBPF* RSF* FSF* CSF) #calc MU
            DArray[21] = ROF* SOBPF* RSF* FSF* CSF #calc dose/MU

            pass
        
        if(isPar_Broken):
            continue
        
        #strip characters
        DArray[1] = DArray[1].strip('S')
        DArray[1] = DArray[1].strip('M')

        #find iso, distal, proxup VQA dose value
        #iso
        for row in ws.iter_rows('A130:A641'):
            for cell in row:
                
                if(cell.value == '#N/A' or str(cell.value) == 'None'):
                    isPDD_Broken = True
                    continue
                
                if(str(ws.cell('D130').value) == 'None'):
                    isPDD_Broken = True
                    continue
                
                if(int(ws.cell('D130').value) != int(cell.value)):
                    continue
                
                Pos = str(cell.coordinate)
                Pos = Pos.strip('A')
                Pos = 'B%s' %(Pos)
                
                DArray[16] = ws.cell(Pos).value
                
                pass
            pass
        
        if(isPDD_Broken):
            continue
        
        #distal
        for row in ws.iter_rows('A130:A641'):
            for cell in row:
                
                if(int(ws.cell('D131').value) != int(cell.value)):
                    continue
                
                Pos = str(cell.coordinate)
                Pos = Pos.strip('A')
                Pos = 'B%s' %(Pos)
                
                DArray[17] = ws.cell(Pos).value
                
                pass
            pass

        #proxup
        for row in ws.iter_rows('A130:A641'):
            for cell in row:
                    
                if(int(ws.cell('D132').value) != int(cell.value)):
                    continue
                
                Pos = str(cell.coordinate)
                Pos = Pos.strip('A')
                Pos = 'B%s' %(Pos)
                DArray[18] = ws.cell(Pos).value
                
                pass
            pass

        #fill into 2D-array for output
        Array.append(DArray)
        
        pass
        
    pass

if(isDebug):
    LogIN = open(DebugLog, 'a')
    
    for i in range(len(Array)):
        for j in range(23):
            #Array[i][j] = Array[i][j].encode('utf-8')
            LogIN.write(str(Array[i][j]))
            LogIN.write(' ')
            pass
        LogIN.write('\n')
        pass
            
    LogIN.close()
    pass

#convert data type
Array = numpy.array(Array, dtype = float)

#escape nan
for i in range(len(Array)):
    for j in range(23):
        val = Array[i, j]
        if(math.isnan(val)):
            Array[i, j] = -1.0
            pass
        pass
    pass

#output
#Output = './Data/EdgeScatterParameter.dat'
numpy.savetxt(Output, Array, 
              fmt = '%d %3.1f %3.1f %3.1f %3.1f %3.1f %3.1f %3.1f %+3.1f %+3.1f %+3.1f %+3.1f %3.4f %3.3f %3.3f %3.3f %3.3f %3.3f %3.3f %1.3f %1.4f %1.4f %4.2f', 
              delimiter = ' ')
